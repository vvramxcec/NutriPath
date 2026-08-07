"""
NutriPath — INDB xlsx → food_items seed migration.

Reads the Anuvaad INDB 2024.11 workbook and writes `supabase/migrations/0004_seed_foods.sql`
with one INSERT per chunk of foods. Deterministic (sorted by food_code) and idempotent
(`ON CONFLICT (food_code) DO NOTHING`).

Per-serving columns from the workbook are mapped to typed `food_items` columns.
The full per-100g row is preserved in the `per_100g` jsonb column for provenance.
Where all per-serving values are absent, per-100g values are used as an approximation
and `serving_fallback = true` is set.

Usage:
    python tools/import_indb.py
    python tools/import_indb.py --input tools/data/Anuvaad_INDB_2024.11.xlsx --output supabase/migrations/0004_seed_foods.sql
"""

from __future__ import annotations

import argparse
import json
import math
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required. Run:  python -m pip install -e tools  "
        "(or:  pip install openpyxl)"
    ) from exc

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = ROOT / "tools" / "data" / "Anuvaad_INDB_2024.11.xlsx"
DEFAULT_OUTPUT = ROOT / "supabase" / "migrations" / "0004_seed_foods.sql"

# per-serving INDB column -> food_items typed column
SERVING_MAP = {
    "unit_serving_energy_kcal": "calories_per_serving_kcal",
    "unit_serving_protein_g": "protein_g",
    "unit_serving_carb_g": "carbs_g",
    "unit_serving_fat_g": "fat_g",
    "unit_serving_freesugar_g": "sugar_g",
    "unit_serving_fibre_g": "fiber_g",
    "unit_serving_sodium_mg": "sodium_mg",
    "unit_serving_potassium_mg": "potassium_mg",
    "unit_serving_phosphorus_mg": "phosphorus_mg",
    "unit_serving_calcium_mg": "calcium_mg",
    "unit_serving_sfa_mg": "sfa_g",  # mg -> g (÷1000)
    "unit_serving_cholesterol_mg": "cholesterol_mg",
    "unit_serving_iron_mg": "iron_mg",
    "unit_serving_vitc_mg": "vitc_mg",
    "unit_serving_folate_ug": "folate_ug",
}

# per-100g INDB column -> typed fallback source (same names as SERVING_MAP targets)
PER100_FALLBACK = {
    "energy_kcal": "calories_per_serving_kcal",
    "protein_g": "protein_g",
    "carb_g": "carbs_g",
    "fat_g": "fat_g",
    "freesugar_g": "sugar_g",
    "fibre_g": "fiber_g",
    "sodium_mg": "sodium_mg",
    "potassium_mg": "potassium_mg",
    "phosphorus_mg": "phosphorus_mg",
    "calcium_mg": "calcium_mg",
    "sfa_mg": "sfa_g",  # mg -> g (÷1000)
    "cholesterol_mg": "cholesterol_mg",
    "iron_mg": "iron_mg",
    "vitc_mg": "vitc_mg",
    "folate_ug": "folate_ug",
}

SKIP_FROM_JSON = {
    "food_code",
    "food_name",
    "primarysource",
    "servings_unit",
}

# round to 2 decimals for compact, sensible SQL
def clean_num(value: Any) -> float | None:
    """Coerce a raw INDB cell to a rounded float, or None for garbage/missing."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
    try:
        num = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(num) or math.isinf(num):
        return None
    if num < 0:  # INDB uses -1 for "not analysed"
        return None
    return round(num, 2)


def sql_literal(value: Any) -> str:
    """Render a Python value as a Postgres literal."""
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        return repr(value)
    return "'" + str(value).replace("'", "''") + "'"


def read_rows(path: Path) -> list[dict[str, Any]]:
    """Load the INDB sheet into a list of row dicts (header -> cell)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    # strip column names
    header = [str(h).strip() if h is not None else "" for h in header]
    index = {name: i for i, name in enumerate(header)}
    data: list[dict[str, Any]] = []
    for values in rows:
        row = {name: values[idx] for name, idx in index.items() if idx < len(values)}
        if not any(v is not None for v in row.values()):
            continue
        data.append(row)
    wb.close()
    return data


def build_food_row(raw: dict[str, Any]) -> dict[str, Any]:
    """Map one INDB row to the food_items column set."""
    per_serving: dict[str, float | None] = {}
    for src_col, tgt_col in SERVING_MAP.items():
        per_serving[tgt_col] = clean_num(raw.get(src_col))

    # per-100g provenance (all columns except identity + unit_serving_*)
    per_100g: dict[str, float | None] = {}
    for col, val in raw.items():
        if col in SKIP_FROM_JSON or col.startswith("unit_serving_"):
            continue
        per_100g[col] = clean_num(val)

    # if every per-serving value is missing, fall back to per-100g
    has_serving = any(v is not None for v in per_serving.values())
    serving_fallback = not has_serving

    if serving_fallback:
        for src_col, tgt_col in PER100_FALLBACK.items():
            v = clean_num(raw.get(src_col))
            if tgt_col == "sfa_g" and v is not None:
                v = round(v / 1000.0, 2)  # mg -> g
            per_serving[tgt_col] = v

    # sfa mg -> g in the typed column
    sfa = per_serving.get("sfa_g")
    if sfa is not None and not serving_fallback:
        per_serving["sfa_g"] = round(sfa / 1000.0, 2)

    name = raw.get("food_name")
    if isinstance(name, str):
        name = name.strip()
    source = str(raw.get("primarysource") or "").strip()
    if source not in {"asc_manual", "bfp_manual", "open_source_recipes"}:
        source = "open_source_recipes"

    return {
        "food_code": str(raw.get("food_code") or "").strip(),
        "name": name or None,
        "source": source,
        **per_serving,
        "per_100g": per_100g,
        "serving_fallback": serving_fallback,
    }


def render_values(food: dict[str, Any]) -> str:
    """Render one food as a SQL VALUES tuple."""
    per_100g_json = json.dumps(food["per_100g"], ensure_ascii=False, separators=(",", ":"))
    cols = [
        "food_code", "name", "source",
        "calories_per_serving_kcal", "protein_g", "carbs_g", "fat_g", "sugar_g",
        "fiber_g", "sodium_mg", "potassium_mg", "phosphorus_mg", "calcium_mg",
        "sfa_g", "cholesterol_mg", "iron_mg", "vitc_mg", "folate_ug",
        "per_100g", "serving_fallback",
    ]
    parts = [sql_literal(food.get(c)) for c in cols[:-2]]
    parts.append("'" + per_100g_json.replace("'", "''") + "'::jsonb")
    parts.append(sql_literal(food["serving_fallback"]))
    return "(" + ", ".join(parts) + ")"


def generate_sql(foods: list[dict[str, Any]]) -> str:
    header = f"""-- 0004_seed_foods.sql
-- GENERATED BY tools/import_indb.py — DO NOT EDIT BY HAND.
-- Source: tools/data/Anuvaad_INDB_2024.11.xlsx ({len(foods)} foods)
-- Regenerate:  python tools/import_indb.py

INSERT INTO food_items
  (food_code, name, source,
   calories_per_serving_kcal, protein_g, carbs_g, fat_g, sugar_g,
   fiber_g, sodium_mg, potassium_mg, phosphorus_mg, calcium_mg,
   sfa_g, cholesterol_mg, iron_mg, vitc_mg, folate_ug,
   per_100g, serving_fallback)
VALUES
"""
    chunks = [foods[i:i + 250] for i in range(0, len(foods), 250)]
    body: list[str] = []
    for chunk in chunks:
        body.append(",\n".join(render_values(f) for f in chunk) + "\n")
    footer = "ON CONFLICT (food_code) DO NOTHING;\n"
    return header + "\n".join(body) + footer


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate 0004_seed_foods.sql from the INDB xlsx")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    if not args.input.exists():
        raise SystemExit(f"input not found: {args.input}")

    rows = read_rows(args.input)
    if not rows:
        raise SystemExit("no data rows found in the workbook")

    foods = [build_food_row(r) for r in rows]
    foods = [f for f in foods if f["food_code"]]  # drop rows without a code
    foods.sort(key=lambda f: f["food_code"])

    missing_codes = [f["food_code"] for f in foods if not f["name"]]
    if missing_codes:
        print(f"WARNING: {len(missing_codes)} rows have no food name")

    fallbacks = sum(1 for f in foods if f["serving_fallback"])
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(generate_sql(foods), encoding="utf-8")
    print(f"wrote {args.output} — {len(foods)} foods, {fallbacks} serving-fallback rows")


if __name__ == "__main__":
    main()
